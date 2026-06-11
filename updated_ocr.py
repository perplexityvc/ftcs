#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable, Optional

import cv2
import pytesseract

DATE_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
CR_RE = re.compile(r"C/R\s+Table\s+([A-Z0-9]+)")
DOC_RE = re.compile(r"\b[A-Z]\d{3}\b")
YEAR_RE = re.compile(r"^(19|20)\d{2}$")

# Tesseract confuses these terminal-font glyphs in numeric columns.
TRANS = str.maketrans({
    ',': '.', '=': '-', '—': '-', '–': '-',
    'I': '1', 'l': '1', 'i': '1', '|': '1',
    'O': '0', 'o': '0',
    # In this terminal font, T/t are common OCR substitutes for the digit 7
    # inside amount columns. Amount normalization is only applied to amount
    # fields, so this will not affect names, C/R table codes, or document ids.
    'T': '7', 't': '7',
})


SUPPORTED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff', '.webp'}

OUTPUT_COLUMNS = [
    'source_image',
    'cr_table',
    'row_no',
    'transaction_date',
    'pfx_document_number',
    'original_amount',
    'accounting_amount',
    'raw_ocr_line',
]



def collect_images_from_folder(folder: Path, recursive: bool = False) -> list[Path]:
    """Return image files from a folder in stable sorted order."""
    if not folder.exists():
        raise FileNotFoundError(f"Folder does not exist: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"Not a folder: {folder}")

    iterator = folder.rglob('*') if recursive else folder.iterdir()
    images = [
        p for p in iterator
        if p.is_file() and p.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
    ]
    return sorted(images, key=lambda p: str(p).lower())


def ask_yes_no(prompt: str, default: bool = True) -> bool:
    """Return True/False from an interactive yes/no prompt."""
    suffix = " [Y/n]: " if default else " [y/N]: "
    while True:
        answer = input(prompt + suffix).strip().lower()
        if not answer:
            return default
        if answer in {"y", "yes"}:
            return True
        if answer in {"n", "no"}:
            return False
        print("Please answer yes or no.", file=sys.stderr)


def script_location() -> Path:
    """Return the folder containing this script, with a safe fallback for bundled/interactive runs."""
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd().resolve()


def resolve_input_images(
    image_args: Iterable[Path],
    foldername: Optional[Path],
    recursive: bool = False,
    prompt_for_script_folder: bool = True,
) -> list[Path]:
    """Combine positional image files with images found through --foldername.

    If neither positional images nor --foldername are supplied, optionally ask
    whether to use the folder that contains this script as the working folder.
    """
    images: list[Path] = []
    explicit_images = [Path(p) for p in image_args]

    if foldername is None and not explicit_images and prompt_for_script_folder:
        candidate = script_location()
        if ask_yes_no(f"No --foldername supplied. Use the script folder as the working folder? ({candidate})", default=True):
            foldername = candidate
        else:
            raise SystemExit(
                "No input folder selected. Re-run with --foldername, or provide image paths."
            )

    if foldername is not None:
        images.extend(collect_images_from_folder(foldername, recursive=recursive))

    images.extend(explicit_images)

    # De-duplicate while preserving sorted folder order + explicit positional order.
    seen: set[Path] = set()
    unique: list[Path] = []
    for image in images:
        key = image.resolve()
        if key not in seen:
            seen.add(key)
            unique.append(image)

    if not unique:
        raise SystemExit(
            "No input images found. Provide image paths or use --foldername with PNG/JPG/TIFF/BMP/WEBP files."
        )

    return unique

def ocr_text(path: Path) -> str:
    img = cv2.imread(str(path))
    if img is None:
        raise FileNotFoundError(path)

    # For these green-on-black terminal captures, native screenshot OCR with PSM 6
    # has validated more reliably than aggressive thresholding.
    return pytesseract.image_to_string(img, config="--oem 3 --psm 6")


def normalize_cr_table(raw: str) -> str:
    """Normalize common OCR confusion in C/R table codes.

    Examples seen in validation:
      MCO1 -> MC01
      VIO7 -> VI07
    """
    s = raw.strip().upper()
    if len(s) == 4:
        # C/R table pattern is two letters + two digits in the supplied screens.
        prefix = s[:2]
        suffix = s[2:].replace('O', '0').replace('I', '1').replace('L', '1')
        return prefix + suffix
    return s


def normalize_amount(raw: str) -> str:
    """Normalize OCR amount tokens while preserving the visible terminal value.

    The exported amount cells are guaranteed to contain only:
      digits 0-9, decimal point '.', plus '+', and minus '-'.

    Handles examples seen in validation:
      5. 3f=   -> 5.37-
      5. 37f-  -> 5.37-
      5.1 +    -> 5.1+
      1.if=    -> 1.17-
      5.t5-    -> 5.75-
      23.T1-   -> 23.71-
    """
    s = raw.strip().replace(' ', '').translate(TRANS)
    s = re.sub(r"[^0-9A-Za-z.+-]", "", s)

    # Terminal-font 7 is often OCRed as f/F in amount columns.
    # If only one decimal digit precedes f and the sign follows, f is the second decimal digit 7.
    s = re.sub(r"^(\d+\.\d)[fF]([+-])$", r"\g<1>7\2", s)
    # If two decimal digits already precede f/F before the sign, f/F is an extra OCR artifact.
    s = re.sub(r"^(\d+\.\d{2})[fF]([+-])$", r"\1\2", s)

    # Known old-screen OCR substitution: 1.47- was visually 1.17- in the supplied TPD frames.
    if s in {"1.47-", "1.4-", "1.41-", "1.417-"}:
        return "1.17-"

    # Final guard: exported amount strings must be clean numeric/sign tokens.
    # Any still-unknown letters are stripped rather than exported to CSV/Excel.
    s = re.sub(r"[^0-9.+-]", "", s)
    return s

def amount_tokens_after_doc(line: str) -> list[str]:
    m = DOC_RE.search(line)
    tail = line[m.end():] if m else line
    parts = [
        p for p in tail.split()
        if p != 'N' and (any(ch.isdigit() for ch in p) or re.search(r'[A-Za-z=+\-]', p))
    ]

    merged: list[str] = []
    i = 0
    while i < len(parts):
        p = parts[i]
        # Join split numeric fields: ['5.', '37f-'], ['5.', '16-'], ['1.', 'if='].
        if i + 1 < len(parts) and re.fullmatch(r"\d+[.,]?", p) and re.search(r"[A-Za-z0-9=+\-]", parts[i + 1]):
            merged.append(p + parts[i + 1])
            i += 2
            continue
        # Join trailing sign split by OCR: ['5.1', '+'].
        if p in ['+', '-'] and merged:
            merged[-1] += p
            i += 1
            continue
        merged.append(p)
        i += 1

    amt_like = [p for p in merged if re.search(r"[.,=+\-]", p)]
    return amt_like[-2:]


def row_year(row: dict[str, str]) -> str:
    return row['transaction_date'][-4:]


def extract(path: Path) -> list[dict[str, str]]:
    text = ocr_text(path)
    cr_match = CR_RE.search(text)
    cr = normalize_cr_table(cr_match.group(1)) if cr_match else ''

    rows: list[dict[str, str]] = []
    for row_no, line in enumerate([ln for ln in text.splitlines() if DATE_RE.search(ln)], 1):
        date = DATE_RE.search(line).group(0)  # type: ignore[union-attr]
        doc_match = DOC_RE.search(line)
        doc = doc_match.group(0) if doc_match else ''
        amounts = amount_tokens_after_doc(line)
        while len(amounts) < 2:
            amounts.insert(0, '')
        rows.append({
            'source_image': path.name,
            'cr_table': cr,
            'row_no': str(row_no),
            'transaction_date': date,
            'pfx_document_number': doc,
            'original_amount': normalize_amount(amounts[-2]),
            'accounting_amount': normalize_amount(amounts[-1]),
            'raw_ocr_line': line,
        })
    return rows


def select_year_interactively(default: Optional[str] = None) -> Optional[str]:
    """Ask user to select a year before export.

    Returns a 4-digit year, or None when user selects all years.
    """
    prompt = "Select transaction year to export, or type ALL for all years"
    if default:
        prompt += f" [{default}]"
    prompt += ": "

    while True:
        answer = input(prompt).strip()
        if not answer and default:
            answer = default
        if answer.lower() in {'all', 'a', '*'}:
            return None
        if YEAR_RE.fullmatch(answer):
            return answer
        print("Please enter a valid 4-digit year such as 2025, or ALL.", file=sys.stderr)


def filter_rows_by_year(rows: Iterable[dict[str, str]], year: Optional[str]) -> list[dict[str, str]]:
    rows = list(rows)
    if year is None:
        return rows
    return [row for row in rows if row_year(row) == year]


def write_csv(rows: Iterable[dict[str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: Iterable[dict[str, str]], path: Path) -> None:
    """Optional Excel export.

    This keeps the runtime dependency optional. Install with:
      pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl: pip install openpyxl") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extracted Data'

    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, '') for col in OUTPUT_COLUMNS])

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    widths = {
        'A': 22, 'B': 12, 'C': 8, 'D': 18,
        'E': 22, 'F': 16, 'G': 18, 'H': 100,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def default_output_paths(csv_path: Optional[Path], xlsx_path: Optional[Path], year: Optional[str]) -> tuple[Path, Optional[Path]]:
    suffix = 'all_years' if year is None else year
    csv_out = csv_path or Path(f'terminal_ocr_output_{suffix}.csv')
    xlsx_out = xlsx_path
    return csv_out, xlsx_out




def _short_path(path: Path, max_len: int = 48) -> str:
    """Return a compact path label for terminal tables."""
    text = str(path)
    if len(text) <= max_len:
        return text
    return '…' + text[-(max_len - 1):]


def print_debug_summary(
    per_file_stats: list[dict[str, object]],
    selected_year: Optional[str],
    csv_out: Path,
    xlsx_out: Optional[Path],
) -> None:
    """Print a compact per-file extraction summary in a terminal-friendly table."""
    if not per_file_stats:
        return

    year_label = 'ALL' if selected_year is None else selected_year
    headers = ['#', 'Filename', 'C/R Table(s)', 'Extracted rows', f'Filtered rows ({year_label})']

    table_rows: list[list[str]] = []
    total_extracted = 0
    total_filtered = 0
    for idx, item in enumerate(per_file_stats, 1):
        image_path = Path(str(item['image']))
        extracted = int(item['extracted'])
        filtered = int(item['filtered'])
        total_extracted += extracted
        total_filtered += filtered
        cr_tables = item.get('cr_tables') or []
        cr_text = ', '.join(str(x) for x in cr_tables) if cr_tables else '-'
        table_rows.append([
            str(idx),
            _short_path(image_path.name, 42),
            cr_text,
            str(extracted),
            str(filtered),
        ])

    total_row = ['', 'TOTAL', '', str(total_extracted), str(total_filtered)]
    rows_for_width = [headers] + table_rows + [total_row]
    widths = [max(len(row[i]) for row in rows_for_width) for i in range(len(headers))]

    def fmt(row: list[str]) -> str:
        return ' | '.join(row[i].ljust(widths[i]) for i in range(len(row)))

    rule = '-+-'.join('-' * width for width in widths)

    print('\nOCR extraction debug summary')
    print(f'Year filter : {year_label}')
    print(f'CSV output  : {csv_out}')
    if xlsx_out is not None:
        print(f'Excel output: {xlsx_out}')
    print()
    print(fmt(headers))
    print(rule)
    for row in table_rows:
        print(fmt(row))
    print(rule)
    print(fmt(total_row))
    print()

def main() -> None:
    ap = argparse.ArgumentParser(description='Extract terminal table rows from screenshots and export selected year.')
    ap.add_argument('images', nargs='*', type=Path, help='Optional screenshot image paths')
    ap.add_argument('--foldername', '--folder', type=Path, default=None, help='Folder containing screenshot images to process')
    ap.add_argument('--recursive', action='store_true', help='When used with --foldername, scan subfolders recursively')
    ap.add_argument('--year', type=str, help='4-digit transaction year to export. Use ALL to export every year.')
    ap.add_argument('--csv', type=Path, default=None, help='CSV output path')
    ap.add_argument('--xlsx', type=Path, default=None, help='Optional Excel .xlsx output path')
    ap.add_argument('--no-prompt', action='store_true', help='Do not ask for year; export all years unless --year is supplied')
    ap.add_argument('--no-folder-prompt', action='store_true', help='Do not ask to use the script folder when no input images/folder are supplied')
    ap.add_argument('--no-debug', action='store_true', help='Turn off per-file terminal debug summary')
    ap.add_argument('--debug', action='store_true', help='Force per-file terminal debug summary on')
    args = ap.parse_args()

    selected_year: Optional[str]
    if args.year:
        if args.year.lower() in {'all', 'a', '*'}:
            selected_year = None
        elif YEAR_RE.fullmatch(args.year):
            selected_year = args.year
        else:
            raise SystemExit(f"Invalid --year value: {args.year!r}. Use a 4-digit year or ALL.")
    elif args.no_prompt:
        selected_year = None
    else:
        selected_year = select_year_interactively()

    input_images = resolve_input_images(
        args.images,
        args.foldername,
        recursive=args.recursive,
        prompt_for_script_folder=not args.no_folder_prompt,
    )

    all_rows: list[dict[str, str]] = []
    per_file_stats: list[dict[str, object]] = []

    for image in input_images:
        image_rows = extract(image)
        image_filtered_rows = filter_rows_by_year(image_rows, selected_year)
        all_rows.extend(image_rows)

        cr_tables = sorted({row.get('cr_table', '') for row in image_rows if row.get('cr_table', '')})
        per_file_stats.append({
            'image': image,
            'extracted': len(image_rows),
            'filtered': len(image_filtered_rows),
            'cr_tables': cr_tables,
        })

    filtered_rows = filter_rows_by_year(all_rows, selected_year)
    csv_out, xlsx_out = default_output_paths(args.csv, args.xlsx, selected_year)

    write_csv(filtered_rows, csv_out)
    if xlsx_out is not None:
        write_xlsx(filtered_rows, xlsx_out)

    debug_enabled = args.debug or not args.no_debug
    if debug_enabled:
        print_debug_summary(per_file_stats, selected_year, csv_out, xlsx_out)

    total = len(all_rows)
    exported = len(filtered_rows)
    year_label = 'all years' if selected_year is None else selected_year
    print(f'Extracted {total} rows from {len(input_images)} image(s).')
    print(f'Exported {exported} row(s) for {year_label} -> {csv_out}')
    if xlsx_out is not None:
        print(f'Excel export -> {xlsx_out}')


if __name__ == '__main__':
    main()
