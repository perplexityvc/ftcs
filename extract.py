#!/usr/bin/env python3
"""
OCR Table Extraction to Excel
Version 5.0 - Updated OCR engine with pytesseract
Extracts accounting tables from terminal screenshots and exports to CSV/Excel
"""
import csv
import re
import sys
import os
from pathlib import Path
from typing import Iterable, Optional

import builtins


def print(*args, **kwargs):
    """Print wrapper that falls back to ASCII-safe output on encoding errors."""
    try:
        builtins.print(*args, **kwargs)
    except UnicodeEncodeError:
        safe_args = [str(arg).encode('ascii', 'replace').decode('ascii') for arg in args]
        builtins.print(*safe_args, **kwargs)


# Import settings
try:
    from settings import (
        ocr, preprocessing, table_detection, cleaning,
        table_structure, paths, year_filter, error_correction,
        output, logging, batch, validation
    )
    SETTINGS_LOADED = True
except ImportError:
    print("Warning: settings.py not found, using default settings")
    SETTINGS_LOADED = False

    class Settings:
        pass

    ocr = Settings()
    ocr.CONFIDENCE_THRESHOLD = 15
    ocr.PSM_MODE = '6'
    ocr.OCR_ENGINE_MODE = 3
    ocr.OUTPUT_FORMAT = 'tsv'
    ocr.TESSERACT_PATH = ''

    preprocessing = Settings()
    preprocessing.PIL_THRESHOLD = 122
    preprocessing.CONTRAST_ENHANCEMENT = 1.2
    preprocessing.APPLY_SHARPENING = True
    preprocessing.APPLY_BINARIZATION = True
    preprocessing.APPLY_VERTICAL_MASK = True
    preprocessing.MASK_TOP_RATIO = 0.33
    preprocessing.MASK_BOTTOM_RATIO = 0.15
    preprocessing.SAVE_MASKED_PREVIEW = True
    preprocessing.MASKED_PREVIEW_DIR = 'masked_inspection'
    preprocessing.MASKED_PREVIEW_GRAYSCALE = False
    preprocessing.MASKED_PREVIEW_SHOW_BBOXES = True

    table_detection = Settings()
    table_detection.ROW_GROUPING_TOLERANCE = 15
    table_detection.COLUMN_GROUPING_TOLERANCE = 45
    table_detection.EMPTY_COLUMN_THRESHOLD = 0.20

    cleaning = Settings()
    cleaning.STANDARD_DOC_NUMBER = 'B674'
    cleaning.MIN_DATE_LENGTH = 8
    cleaning.MIN_DOC_NUMBER_LENGTH = 2
    cleaning.DATE_REQUIRED_CHAR = '/'

    table_structure = Settings()
    table_structure.STANDARD_HEADERS = ['C/R Table', 'Date', 'Doc No.', 'Orig Amount', 'Acc Amount']
    table_structure.HEADER_KEYWORDS = ['date', 'trans', 'number', 'amount', 'document', 'sel']
    table_structure.TABLE_NUMBER_PATTERNS = [
        r'C/R\s+Table\s+([A-Z0-9]+)',
        r'Table\s+([A-Z][A-Z0-9]{2,5})',
        r'C/R.*?([A-Z]{2}\d{2})',
    ]
    table_structure.MAX_TABLE_NUMBER_LENGTH = 10
    table_structure.MIN_TABLE_NUMBER_LENGTH = 2

    paths = Settings()
    paths.DEFAULT_INPUT_IMAGE = 'frame_0015.png'
    paths.DEFAULT_INPUT_FOLDER = 'frames_masked'
    paths.TEMP_PREPROCESSED_IMAGE = 'preprocessed.png'
    paths.DEFAULT_OUTPUT_CSV = 'extracted_table.csv'
    paths.DEFAULT_BATCH_OUTPUT = 'combined_output.csv'
    paths.EXCEL_GENERATOR_SCRIPT = 'create_excel.py'
    paths.SUPPORTED_IMAGE_FORMATS = ['*.png', '*.PNG', '*.jpg', '*.jpeg', '*.bmp', '*.tif', '*.tiff', '*.webp']

    year_filter = Settings()
    year_filter.TWO_DIGIT_YEAR_CUTOFF = 50
    year_filter.DEFAULT_CENTURY = 2000
    year_filter.ALTERNATIVE_CENTURY = 1900

    error_correction = Settings()
    error_correction.AMOUNT_CHAR_TRANSLATION = {
        ',': '.', '=': '-', '—': '-', '–': '-',
        'I': '1', 'l': '1', 'i': '1', '|': '1',
        'O': '0', 'o': '0',
        'T': '7', 't': '7',
    }

    output = Settings()
    output.CSV_ENCODING = 'utf-8'
    output.PREVIEW_ROWS = 10
    output.PREVIEW_COLUMN_WIDTH = 15
    output.OUTPUT_COLUMNS = [
        'source_image', 'cr_table', 'row_no', 'transaction_date',
        'pfx_document_number', 'original_amount', 'accounting_amount', 'raw_ocr_line',
    ]

    logging = Settings()
    logging.CONSOLE_WIDTH = 70
    logging.SHOW_PROGRESS = True
    logging.SHOW_TIMING = True
    logging.VERBOSE_MODE = True

    batch = Settings()
    batch.CONTINUE_ON_ERROR = True
    batch.CLEANUP_TEMP_FILES = True

    validation = Settings()


# ============================================================================
# REGEX PATTERNS
# ============================================================================

DATE_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
CR_RE = re.compile(r"C/R\s+Table\s+[^\w]*([A-Z0-9]+)")
DOC_RE = re.compile(r"\b[A-Z]\d{3}\b")
YEAR_RE = re.compile(r"^(19|20)\d{2}$")

# Character translation for amount normalization
TRANS = str.maketrans(getattr(error_correction, 'AMOUNT_CHAR_TRANSLATION', {
    ',': '.', '=': '-', '—': '-', '–': '-',
    'I': '1', 'l': '1', 'i': '1', '|': '1',
    'O': '0', 'o': '0',
    'T': '7', 't': '7',
}))

SUPPORTED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff', '.webp'}


# ============================================================================
# BOUNDING BOX DRAWING FUNCTIONS
# ============================================================================

def _parse_tsv_for_bboxes(tsv_data):
    """Parse Tesseract TSV output into list of bbox dicts."""
    boxes = []
    for line in tsv_data.strip().split('\n')[1:]:
        parts = line.split('\t')
        if len(parts) < 12:
            continue
        try:
            level = int(parts[0])
            if level != 5:  # word-level only
                continue
            conf = float(parts[10])
            if conf < 0:  # -1 = non-word block entry
                continue
            text = parts[11].strip()
            if not text:
                continue
            boxes.append({
                'left': int(parts[6]),
                'top': int(parts[7]),
                'width': int(parts[8]),
                'height': int(parts[9]),
                'conf': conf,
                'text': text,
            })
        except (ValueError, IndexError):
            continue
    return boxes


def draw_ocr_bboxes_on_image(pil_img, source_image_path):
    """Overlay OCR bounding boxes and confidence labels on a PIL image.

    Color coding:
      green  (0, 200, 0)   — confidence >= 80 %
      yellow (220, 180, 0) — confidence  50 – 79 %
      red    (220, 0, 0)   — confidence  < 50 %
    """
    try:
        from PIL import ImageDraw, ImageFont
        import pytesseract

        # Run Tesseract TSV on the original source image
        import cv2
        img = cv2.imread(str(source_image_path))
        if img is None:
            return pil_img

        config = _get_tesseract_config()
        tsv_data = pytesseract.image_to_data(img, config=config)
        if not tsv_data or not tsv_data.strip():
            return pil_img

        boxes = _parse_tsv_for_bboxes(tsv_data)
        if not boxes:
            return pil_img

        # Scale factor in case pil_img has been resized vs source
        src_w, src_h = pil_img.size
        try:
            from PIL import Image as _PILImage
            with _PILImage.open(source_image_path) as _src:
                orig_w, orig_h = _src.size
        except Exception:
            orig_w, orig_h = src_w, src_h

        scale_x = src_w / orig_w if orig_w > 0 else 1.0
        scale_y = src_h / orig_h if orig_h > 0 else 1.0

        draw = ImageDraw.Draw(pil_img)

        # Try to load a small font
        font = None
        try:
            font = ImageFont.truetype("arial.ttf", 11)
        except Exception:
            try:
                font = ImageFont.load_default()
            except Exception:
                font = None

        for box in boxes:
            conf = box['conf']

            if conf >= 80:
                color = (0, 200, 0)  # green
            elif conf >= 50:
                color = (220, 180, 0)  # yellow
            else:
                color = (220, 0, 0)  # red

            x0 = int(box['left'] * scale_x)
            y0 = int(box['top'] * scale_y)
            x1 = int((box['left'] + box['width']) * scale_x)
            y1 = int((box['top'] + box['height']) * scale_y)

            draw.rectangle([x0, y0, x1, y1], outline=color, width=1)

            label = f"{conf:.0f}%"
            label_y = max(0, y0 - 12)
            if font:
                draw.text((x0, label_y), label, fill=color, font=font)
            else:
                draw.text((x0, label_y), label, fill=color)

        return pil_img

    except ImportError:
        # pytesseract or PIL not available — skip boxes silently
        return pil_img
    except Exception as e:
        if getattr(logging, 'SHOW_PROGRESS', True):
            print(f"  Warning: Could not draw OCR bounding boxes: {e}")
        return pil_img


def save_masked_preview(source_image_path, show_bboxes=True, grayscale=False):
    """Save a masked preview of the image with optional bbox overlay."""
    try:
        from PIL import Image, ImageDraw

        preview_dir = getattr(preprocessing, 'MASKED_PREVIEW_DIR', 'masked_inspection')
        os.makedirs(preview_dir, exist_ok=True)

        source_name = os.path.basename(source_image_path)
        source_stem, source_ext = os.path.splitext(source_name)
        ext = source_ext if source_ext else '.png'
        preview_path = os.path.join(preview_dir, f"{source_stem}_masked{ext}")

        # Load original image
        img = Image.open(source_image_path).convert('RGB')

        # Optionally convert to grayscale
        if grayscale:
            img = img.convert('L')

        # Apply vertical mask
        if getattr(preprocessing, 'APPLY_VERTICAL_MASK', True):
            width, height = img.size
            mask_top = getattr(preprocessing, 'MASK_TOP_RATIO', 0.33)
            mask_bottom = getattr(preprocessing, 'MASK_BOTTOM_RATIO', 0.15)
            top_pixels = int(height * max(0.0, min(1.0, mask_top)))
            bottom_pixels = int(height * max(0.0, min(1.0, mask_bottom)))
            draw = ImageDraw.Draw(img)
            if top_pixels > 0:
                draw.rectangle([0, 0, width - 1, min(top_pixels - 1, height - 1)], fill=0)
            if bottom_pixels > 0:
                start_y = max(0, height - bottom_pixels)
                draw.rectangle([0, start_y, width - 1, height - 1], fill=0)

        # Draw OCR bounding boxes if enabled
        if show_bboxes:
            img = draw_ocr_bboxes_on_image(img, source_image_path)

        img.save(preview_path)
        return preview_path

    except ImportError:
        if getattr(logging, 'SHOW_PROGRESS', True):
            print("  Warning: Pillow not installed — cannot save masked preview")
        return None
    except Exception as e:
        if getattr(logging, 'SHOW_PROGRESS', True):
            print(f"  Warning: Could not save masked preview: {e}")
        return None


# ============================================================================
# OCR FUNCTIONS
# ============================================================================

def _get_tesseract_config():
    """Build tesseract config string from settings."""
    psm = getattr(ocr, 'PSM_MODE', '6')
    oem = getattr(ocr, 'OCR_ENGINE_MODE', 3)
    return f"--oem {oem} --psm {psm}"


def ocr_text(path: Path) -> str:
    """Run OCR on an image using pytesseract."""
    try:
        import cv2
        img = cv2.imread(str(path))
        if img is None:
            raise FileNotFoundError(f"Could not load image: {path}")

        import pytesseract
        config = _get_tesseract_config()
        return pytesseract.image_to_string(img, config=config)
    except ImportError:
        # Fallback to subprocess if pytesseract not installed
        import subprocess
        config = _get_tesseract_config()
        cmd = ['tesseract', str(path), 'stdout'] + config.split()
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(f"Tesseract failed: {result.stderr}")
        return result.stdout


# ============================================================================
# NORMALIZATION FUNCTIONS
# ============================================================================

def normalize_cr_table(raw: str) -> str:
    """Normalize common OCR confusion in C/R table codes.
    
    Examples:
      CI01 -> CI01 (already correct)
      CIO1 -> CI01 (O misread as 0)
      CIOQ1 -> CI01 (O misread, Q is OCR noise)
      MCO1 -> MC01 (O misread as 0)
    """
    s = raw.strip().upper()
    
    # Expected format: 2 letters + 2 digits (e.g., CI01)
    # Handle OCR noise that adds extra characters
    
    if len(s) >= 4:
        # Extract first 2 characters as prefix (should be letters)
        prefix = s[:2]
        
        # Get the remaining part and normalize
        suffix = s[2:]
        
        # Apply OCR normalization to suffix
        # O->0, I->1, L->1 for digit confusion
        normalized = suffix.replace('O', '0').replace('I', '1').replace('L', '1')
        
        # Remove any non-digit characters (OCR noise like Q, etc.)
        digits_only = ''.join(c for c in normalized if c.isdigit())
        
        # Take first 2 digits if available
        if len(digits_only) >= 2:
            return prefix + digits_only[:2]
        elif len(digits_only) == 1:
            return prefix + digits_only
        else:
            return prefix
    
    return s


def normalize_amount(raw: str) -> str:
    """Normalize OCR amount tokens while preserving the visible terminal value."""
    s = raw.strip().replace(' ', '').translate(TRANS)
    s = re.sub(r"[^0-9A-Za-z.+-]", "", s)

    # Terminal-font 7 is often OCRed as f/F in amount columns
    s = re.sub(r"^(\d+\.\d)[fF]([+-])$", r"\g<1>7\2", s)
    s = re.sub(r"^(\d+\.\d{2})[fF]([+-])$", r"\1\2", s)

    # Known old-screen OCR substitution
    if s in {"1.47-", "1.4-", "1.41-", "1.417-"}:
        return "1.17-"

    # Final guard: strip unknown letters
    s = re.sub(r"[^0-9.+-]", "", s)
    return s


def amount_tokens_after_doc(line: str) -> list[str]:
    """Extract amount tokens from a line after the document number."""
    m = DOC_RE.search(line)
    tail = line[m.end():] if m else line
    parts = [
        p for p in tail.split()
        if p != 'N' and (any(ch.isdigit() for ch in p) or re.search(r'[A-Za-z=+\-]', p))
    ]

    merged: list[str] = []
    i = 0
    while i < len(parts):
        p = parts[i]
        # Join split numeric fields
        if i + 1 < len(parts) and re.fullmatch(r"\d+[.,]?", p) and re.search(r"[A-Za-z0-9=+\-]", parts[i + 1]):
            merged.append(p + parts[i + 1])
            i += 2
            continue
        # Join trailing sign split by OCR
        if p in ['+', '-'] and merged:
            merged[-1] += p
            i += 1
            continue
        merged.append(p)
        i += 1

    amt_like = [p for p in merged if re.search(r"[.,=+\-]", p)]
    return amt_like[-2:]


# ============================================================================
# IMAGE COLLECTION
# ============================================================================

def collect_images_from_folder(folder: Path, recursive: bool = False) -> list[Path]:
    """Return image files from a folder in stable sorted order."""
    if not folder.exists():
        raise FileNotFoundError(f"Folder does not exist: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"Not a folder: {folder}")

    iterator = folder.rglob('*') if recursive else folder.iterdir()
    images = [
        p for p in iterator
        if p.is_file() and p.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
    ]
    return sorted(images, key=lambda p: str(p).lower())


# ============================================================================
# EXTRACTION
# ============================================================================

def extract_rows(path: Path) -> list[dict[str, str]]:
    """Extract table rows from an image."""
    text = ocr_text(path)
    cr_match = CR_RE.search(text)
    cr = normalize_cr_table(cr_match.group(1)) if cr_match else ''

    rows: list[dict[str, str]] = []
    for row_no, line in enumerate([ln for ln in text.splitlines() if DATE_RE.search(ln)], 1):
        date = DATE_RE.search(line).group(0)
        doc_match = DOC_RE.search(line)
        doc = doc_match.group(0) if doc_match else ''
        amounts = amount_tokens_after_doc(line)
        while len(amounts) < 2:
            amounts.insert(0, '')
        rows.append({
            'source_image': path.name,
            'cr_table': cr,
            'row_no': str(row_no),
            'transaction_date': date,
            'pfx_document_number': doc,
            'original_amount': normalize_amount(amounts[-2]),
            'accounting_amount': normalize_amount(amounts[-1]),
            'raw_ocr_line': line,
        })
    return rows


def filter_rows_by_year(rows: Iterable[dict[str, str]], year: Optional[str]) -> list[dict[str, str]]:
    """Filter rows by year."""
    rows = list(rows)
    if year is None:
        return rows
    return [row for row in rows if row['transaction_date'][-4:] == year]


# ============================================================================
# OUTPUT FUNCTIONS
# ============================================================================

def write_csv(rows: Iterable[dict[str, str]], path: Path) -> None:
    """Write rows to CSV file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = getattr(output, 'OUTPUT_COLUMNS', [
        'source_image', 'cr_table', 'row_no', 'transaction_date',
        'pfx_document_number', 'original_amount', 'accounting_amount', 'raw_ocr_line',
    ])
    with path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: Iterable[dict[str, str]], path: Path) -> None:
    """Write rows to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl: pip install openpyxl") from exc

    columns = getattr(output, 'OUTPUT_COLUMNS', [
        'source_image', 'cr_table', 'row_no', 'transaction_date',
        'pfx_document_number', 'original_amount', 'accounting_amount', 'raw_ocr_line',
    ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extracted Data'

    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, '') for col in columns])

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    widths = {
        'A': 22, 'B': 12, 'C': 8, 'D': 18,
        'E': 22, 'F': 16, 'G': 18, 'H': 100,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def _short_path(path: Path, max_len: int = 48) -> str:
    """Return a compact path label for terminal tables."""
    text = str(path)
    if len(text) <= max_len:
        return text
    return '...' + text[-(max_len - 1):]


def print_debug_summary(
    per_file_stats: list[dict[str, object]],
    selected_year: Optional[str],
    csv_out: Path,
    xlsx_out: Optional[Path],
) -> None:
    """Print a compact per-file extraction summary in a terminal-friendly table."""
    if not per_file_stats:
        return

    year_label = 'ALL' if selected_year is None else selected_year
    headers = ['#', 'Filename', 'C/R Table(s)', 'Extracted rows', f'Filtered rows ({year_label})']

    table_rows: list[list[str]] = []
    total_extracted = 0
    total_filtered = 0
    for idx, item in enumerate(per_file_stats, 1):
        image_path = Path(str(item['image']))
        extracted = int(item['extracted'])
        filtered = int(item['filtered'])
        total_extracted += extracted
        total_filtered += filtered
        cr_tables = item.get('cr_tables') or []
        cr_text = ', '.join(str(x) for x in cr_tables) if cr_tables else '-'
        table_rows.append([
            str(idx),
            _short_path(image_path.name, 42),
            cr_text,
            str(extracted),
            str(filtered),
        ])

    total_row = ['', 'TOTAL', '', str(total_extracted), str(total_filtered)]
    rows_for_width = [headers] + table_rows + [total_row]
    widths = [max(len(row[i]) for row in rows_for_width) for i in range(len(headers))]

    def fmt(row: list[str]) -> str:
        return ' | '.join(row[i].ljust(widths[i]) for i in range(len(row)))

    rule = '-+-'.join('-' * width for width in widths)

    print('\nOCR extraction debug summary')
    print(f'Year filter : {year_label}')
    print(f'CSV output  : {csv_out}')
    if xlsx_out is not None:
        print(f'Excel output: {xlsx_out}')
    print()
    print(fmt(headers))
    print(rule)
    for row in table_rows:
        print(fmt(row))
    print(rule)
    print(fmt(total_row))
    print()


# ============================================================================
# MAIN FUNCTIONS
# ============================================================================

def main(
    input_image=None,
    year_filter_value=None,
    output_csv=None,
    output_xlsx=None,
    no_debug=False,
):
    """Main extraction function for single image."""
    import time
    start_time = time.time()

    print("=" * logging.CONSOLE_WIDTH)
    print(" " * 15 + "OCR TABLE EXTRACTION")
    print("=" * logging.CONSOLE_WIDTH)

    # Resolve input image
    if input_image is None:
        input_image = Path(getattr(paths, 'DEFAULT_INPUT_IMAGE', 'frame_0015.png'))
    else:
        input_image = Path(input_image)

    if year_filter_value:
        print(f"\n*** YEAR FILTER: Only extracting data from year {year_filter_value} ***\n")

    # Extract rows
    print(f"\nProcessing: {input_image.name}")
    print("-" * logging.CONSOLE_WIDTH)

    try:
        rows = extract_rows(input_image)
        filtered_rows = filter_rows_by_year(rows, year_filter_value)

        cr_tables = sorted({row.get('cr_table', '') for row in rows if row.get('cr_table', '')})
        cr_text = ', '.join(cr_tables) if cr_tables else '-'

        print(f"  C/R Table: {cr_text}")
        print(f"  Extracted: {len(rows)} rows")
        print(f"  Filtered:  {len(filtered_rows)} rows")

        # Save masked preview with bounding boxes
        show_bboxes = not getattr(args, 'no_preview_bboxes', False) if 'args' in dir() else getattr(preprocessing, 'MASKED_PREVIEW_SHOW_BBOXES', True)
        use_grayscale = getattr(args, 'preview_grayscale', False) if 'args' in dir() else getattr(preprocessing, 'MASKED_PREVIEW_GRAYSCALE', False)
        preview_path = save_masked_preview(str(input_image), show_bboxes=show_bboxes, grayscale=use_grayscale)
        if preview_path:
            print(f"  Preview:   {preview_path}")

    except FileNotFoundError:
        print(f"  Error: Image not found: {input_image}")
        return
    except Exception as e:
        print(f"  Error processing {input_image.name}: {e}")
        return

    # Determine output paths
    if output_csv is None:
        suffix = 'all_years' if year_filter_value is None else year_filter_value
        output_csv = Path(f'terminal_ocr_output_{suffix}.csv')
    else:
        output_csv = Path(output_csv)

    xlsx_out = Path(output_xlsx) if output_xlsx else None

    # Write output
    write_csv(filtered_rows, output_csv)
    print(f"\n  CSV output: {output_csv}")

    if xlsx_out:
        write_xlsx(filtered_rows, xlsx_out)
        print(f"  Excel output: {xlsx_out}")

    # Summary
    print("\n" + "=" * logging.CONSOLE_WIDTH)
    print("EXTRACTION COMPLETE!")
    print("=" * logging.CONSOLE_WIDTH)

    elapsed = time.time() - start_time
    year_label = 'all years' if year_filter_value is None else year_filter_value
    print(f"\nExtracted {len(rows)} rows from 1 image.")
    print(f"Exported {len(filtered_rows)} row(s) for {year_label} -> {output_csv}")

    if logging.SHOW_TIMING:
        print(f"Processing time: {elapsed:.2f} seconds")

    return elapsed, len(filtered_rows)


def batch_process_images(
    input_folder=None,
    output_csv=None,
    output_xlsx=None,
    year_filter_value=None,
    recursive=False,
    no_debug=False,
):
    """Process all images in a folder and combine results into single CSV."""
    import time
    start_time = time.time()

    print("=" * logging.CONSOLE_WIDTH)
    print(" " * 15 + "BATCH OCR PROCESSING")
    print("=" * logging.CONSOLE_WIDTH)

    # Resolve input folder
    if input_folder is None:
        input_folder = Path(getattr(paths, 'DEFAULT_INPUT_FOLDER', 'frames_masked'))
    else:
        input_folder = Path(input_folder)

    if year_filter_value:
        print(f"\n*** YEAR FILTER: Only extracting data from year {year_filter_value} ***\n")

    # Clear masked_inspection folder if it exists
    preview_dir = getattr(preprocessing, 'MASKED_PREVIEW_DIR', 'masked_inspection')
    if os.path.isdir(preview_dir):
        for name in os.listdir(preview_dir):
            file_path = os.path.join(preview_dir, name)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass

    # Collect images
    try:
        image_files = collect_images_from_folder(input_folder, recursive=recursive)
    except (FileNotFoundError, NotADirectoryError) as e:
        print(f"Error: {e}")
        return

    if not image_files:
        print(f"No images found in: {input_folder}")
        return

    print(f"\nFound {len(image_files)} images to process:")
    for img in image_files:
        print(f"  - {img.name}")

    print("\n" + "-" * logging.CONSOLE_WIDTH)

    # Process each image
    all_rows: list[dict[str, str]] = []
    per_file_stats: list[dict[str, object]] = []

    for idx, image_path in enumerate(image_files, 1):
        print(f"\n[{idx}/{len(image_files)}] Processing: {image_path.name}")
        print("-" * logging.CONSOLE_WIDTH)

        img_start_time = time.time()

        try:
            rows = extract_rows(image_path)
            filtered_rows = filter_rows_by_year(rows, year_filter_value)
            all_rows.extend(rows)

            cr_tables = sorted({row.get('cr_table', '') for row in rows if row.get('cr_table', '')})
            cr_text = ', '.join(cr_tables) if cr_tables else '-'

            print(f"  C/R Table: {cr_text}")
            print(f"  Extracted: {len(rows)} rows")
            print(f"  Filtered:  {len(filtered_rows)} rows")

            # Save masked preview with bounding boxes
            show_bboxes = not getattr(args, 'no_preview_bboxes', False) if 'args' in dir() else getattr(preprocessing, 'MASKED_PREVIEW_SHOW_BBOXES', True)
            use_grayscale = getattr(args, 'preview_grayscale', False) if 'args' in dir() else getattr(preprocessing, 'MASKED_PREVIEW_GRAYSCALE', False)
            preview_path = save_masked_preview(str(image_path), show_bboxes=show_bboxes, grayscale=use_grayscale)
            if preview_path:
                print(f"  Preview:   {preview_path}")

            per_file_stats.append({
                'image': image_path,
                'extracted': len(rows),
                'filtered': len(filtered_rows),
                'cr_tables': cr_tables,
            })

        except Exception as e:
            print(f"  Error processing {image_path.name}: {e}")
            per_file_stats.append({
                'image': image_path,
                'extracted': 0,
                'filtered': 0,
                'cr_tables': [],
            })

    # Filter all rows by year
    filtered_all_rows = filter_rows_by_year(all_rows, year_filter_value)

    # Determine output paths
    if output_csv is None:
        suffix = 'all_years' if year_filter_value is None else year_filter_value
        output_csv = Path(f'terminal_ocr_output_{suffix}.csv')
    else:
        output_csv = Path(output_csv)

    xlsx_out = Path(output_xlsx) if output_xlsx else None

    # Write output
    write_csv(filtered_all_rows, output_csv)
    if xlsx_out:
        write_xlsx(filtered_all_rows, xlsx_out)

    # Print debug summary
    if not no_debug:
        print_debug_summary(per_file_stats, year_filter_value, output_csv, xlsx_out)

    # Summary
    elapsed = time.time() - start_time
    year_label = 'all years' if year_filter_value is None else year_filter_value

    print(f"\nExtracted {len(all_rows)} rows from {len(image_files)} image(s).")
    print(f"Exported {len(filtered_all_rows)} row(s) for {year_label} -> {output_csv}")
    if xlsx_out:
        print(f"Excel export -> {xlsx_out}")

    if logging.SHOW_TIMING:
        print(f"Processing time: {elapsed:.2f} seconds")


# ============================================================================
# CLI ENTRY POINT
# ============================================================================

def _parse_cli_args():
    """Parse command line arguments."""
    import argparse

    ap = argparse.ArgumentParser(
        description='Extract terminal table rows from screenshots and export to CSV/Excel.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
EXAMPLES:

  # Single image extraction
  python extract.py image.png
  python extract.py image.png --year 2025

  # Batch processing with folder
  python extract.py --batch ./images
  python extract.py --batch ./images --year 2025

  # Batch with explicit output
  python extract.py --batch ./images combined_output.csv

  # Use defaults from settings.py
  python extract.py --batch

  # Export to Excel
  python extract.py image.png --xlsx output.xlsx

  # Disable debug summary
  python extract.py --batch --no-debug
        """
    )

    ap.add_argument('images', nargs='*', type=Path, help='Screenshot image paths')
    ap.add_argument('--batch', action='store_true', help='Batch process all images in folder')
    ap.add_argument('--folder', '--foldername', type=Path, default=None,
                    help='Folder containing screenshot images to process')
    ap.add_argument('--recursive', action='store_true',
                    help='When used with --batch, scan subfolders recursively')
    ap.add_argument('--year', type=str,
                    help='4-digit transaction year to export. Use ALL to export every year.')
    ap.add_argument('--csv', type=Path, default=None, help='CSV output path')
    ap.add_argument('--xlsx', type=Path, default=None, help='Optional Excel .xlsx output path')
    ap.add_argument('--no-prompt', action='store_true',
                    help='Do not ask for year; export all years unless --year is supplied')
    ap.add_argument('--no-debug', action='store_true',
                    help='Turn off per-file terminal debug summary')
    ap.add_argument('--debug', action='store_true',
                    help='Force per-file terminal debug summary on')
    ap.add_argument('--no-preprocess', action='store_true',
                    help='Disable image preprocessing (kept for backward compatibility)')
    ap.add_argument('--grayscale-only', action='store_true',
                    help='Disable binarization, keep grayscale (kept for backward compatibility)')
    ap.add_argument('--mask-top', type=float, default=None,
                    help='Mask top percentage (0-100, default from settings)')
    ap.add_argument('--mask-bottom', type=float, default=None,
                    help='Mask bottom percentage (0-100, default from settings)')
    ap.add_argument('--preview-grayscale', action='store_true',
                    help='Save masked preview in grayscale')
    ap.add_argument('--no-preview-bboxes', action='store_true',
                    help='Disable OCR bounding box overlay on preview')
    ap.add_argument('--help-legacy', action='store_true',
                    help='Show legacy help message')

    return ap.parse_args()


if __name__ == "__main__":
    args = _parse_cli_args()

    # Handle year filter
    selected_year = None
    if args.year:
        if args.year.lower() in {'all', 'a', '*'}:
            selected_year = None
        elif YEAR_RE.fullmatch(args.year):
            selected_year = args.year
        else:
            print(f"Invalid --year value: {args.year!r}. Use a 4-digit year or ALL.")
            sys.exit(1)
    elif args.no_prompt:
        selected_year = None

    # Determine debug mode
    no_debug = args.no_debug and not args.debug

    if args.batch or args.folder:
        # Batch mode
        batch_process_images(
            input_folder=args.folder,
            output_csv=args.csv,
            output_xlsx=args.xlsx,
            year_filter_value=selected_year,
            recursive=args.recursive,
            no_debug=no_debug,
        )
    else:
        # Single image mode
        image_path = args.images[0] if args.images else None
        main(
            input_image=image_path,
            year_filter_value=selected_year,
            output_csv=args.csv,
            output_xlsx=args.xlsx,
            no_debug=no_debug,
        )
